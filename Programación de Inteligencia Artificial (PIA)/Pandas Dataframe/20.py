import pandas as pd
from openpyxl import load_workbook

# Crear un DataFrame de ejemplo (puedes usar el tuyo propio)
data = {
    'Nombre': ['Alejandro', 'María', 'Carlos'],
    'Programación T1': [9.654, 7.345, 5.123],
    'Base de Datos T1': [8.112, 6.789, 9.456],
    'Lenguajes T1': [9.999, 6.321, 7.654]
}
df = pd.DataFrame(data)

# Guardar el DataFrame a un archivo Excel
ruta_archivo_excel = 'alumnos_notas.xlsx'
df.to_excel(ruta_archivo_excel, index=False)

# Cargar el archivo Excel para formatearlo
wb = load_workbook(ruta_archivo_excel)
ws = wb.active

# Formatear las celdas de las notas para que tengan 2 decimales
# Suponiendo que las notas están en las columnas de 'Programación T1', 'Base de Datos T1', 'Lenguajes T1', etc.
columnas_notas = ['Programación T1', 'Base de Datos T1', 'Lenguajes T1']

# Buscar las columnas de notas en el DataFrame y aplicar el formato a las celdas correspondientes
for col_name in columnas_notas:
    col_idx = df.columns.get_loc(col_name) + 1  # Obtenemos el índice de la columna en Excel
    for row in range(2, len(df) + 2):  # Empezamos desde la fila 2 (para saltar el encabezado)
        cell = ws.cell(row=row, column=col_idx)
        cell.number_format = '0.00'  # Establecemos el formato a dos decimales

# Guardar el archivo Excel con el formato aplicado
wb.save('alumnos_notas_formateado.xlsx')

print("Archivo Excel con formato de dos decimales guardado en: alumnos_notas_formateado.xlsx")
